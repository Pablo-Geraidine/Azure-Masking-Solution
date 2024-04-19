# Databricks notebook source
# MAGIC %pip install openpyxl faker azure-identity -q

# COMMAND ----------

pip install azure-storage-blob

# COMMAND ----------

from pyspark.sql import SparkSession
import pandas as pd
import os
import uuid
import string
import random
import time
import re
from datetime import datetime
from faker import Faker
import shutil
from shutil import copyfile
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

# COMMAND ----------

start_runtime = time.time()

# COMMAND ----------

configs = {
  "fs.azure.account.auth.type": "OAuth",
  "fs.azure.account.oauth.provider.type": "org.apache.hadoop.fs.azurebfs.oauth2.ClientCredsTokenProvider",
  "fs.azure.account.oauth2.client.id": "f53acde7-87dc-4744-b03b-05b298a79ab6",
  "fs.azure.account.oauth2.client.secret": "bjx8Q~z5o7RL3F49WF4ej.quAj9Gdyn0ECmVgcjn", 
  "fs.azure.account.oauth2.client.endpoint": f"https://login.microsoftonline.com/8f72f2f2-1c4b-4d6a-a13a-3825abea5a3c/oauth2/token",
  "fs.azure.createRemoteFileSystemDuringInitialization": "true"
}

# COMMAND ----------

# Checks if mount point exists; if not, create it
if any(mount.mountPoint == '/mnt/masking_config' for mount in dbutils.fs.mounts()):
  mount_point = '/mnt/masking_config'
  print('mount point already exists')
else:
  dbutils.fs.mount(
    source = "abfss://masking-config@pabloconfig.dfs.core.windows.net/",
    mount_point = "/mnt/masking_config",
    extra_configs = configs)
  print('datalake mount done')

# COMMAND ----------

#dbutils.widgets.removeAll()
dbutils.widgets.dropdown("Project", "Project Intake", ['Project Intake', 'Test Project 1'])

# COMMAND ----------

#This cell defines paths for the unique masking definition files (intake form) and mask dictionaries (lookup table), grouped by project.

project_definition = dbutils.widgets.get("Project")
print(f"Project: {project_definition}")
mount_point_path = '/dbfs/mnt/masking_config/Project Intake/'


#Enter missing locations once ADLS2 account created
masking_rules_location = '/dbfs/mnt/masking_config/Project Intake/Configurable Masking Rulebook - blank -macro - enhanced.xlsm'
#masking_rules_location_no_macro = '/dbfs/mnt/masking_config/Project Intake/Configurable Masking Rulebook - blank.xlsx'
new_excel_path = '/dbfs/mnt/masking_config/Project Intake/Configurable Masking Rulebook Modified.xlsm'
#new_excel_path_no_macro = '/dbfs/mnt/masking_config/Project Intake/Configurable Masking Rulebook Modified.xlsx'
local_path = "/tmp/Configurable Masking Rulebook Modified.xlsm"
#copyfile(masking_rules_location, new_excel_path)
#masking_dictionary_location = '/dbfs/mnt/masking_config/Project Intake/mask_dictionary.csv'
if project_definition == 'Test Project 1':
  configurable_masking_rulebook = f'/dbfs/mnt/masking_config/Test Project 1/{project_definition} Configurable Masking Rulebook Template.xlsm'
  reference_filename_mapping = f'/dbfs/mnt/masking_config/Test Project 1/{project_definition} Reference Filename Mapping.csv'
elif project_definition == 'Test Project 2':
  #Enter missing locations once ADLS2 account created
  masking_rules_location = ''
  masking_dictionary_location = ''

Masking_definition_file = masking_rules_location
#mask_map_table_path = masking_dictionary_location

# COMMAND ----------

# Configure ADLS connection to input location
df_read_location = pd.read_excel(Masking_definition_file, engine='openpyxl', sheet_name='Read Location')
display(df_read_location)

read_storage_account = df_read_location['Storage Account'][0]
read_container = df_read_location['Container'][0]
read_directory_path = df_read_location['Directory Path'][0]
if str(read_directory_path) == 'nan':
    read_directory_path = ''
print(f'Read storage account: {read_storage_account}')
print(f'Read container: {read_container}')
print(f'Read directory path: {read_directory_path}')

# Checks if mount point exists; if not, create it
input_mount_point_name = ('/mnt/input_path_' + read_storage_account + '_' + read_container)
if any(mount.mountPoint == input_mount_point_name for mount in dbutils.fs.mounts()):
  mount_point = input_mount_point_name
  print('mount point already exists')
else:
  dbutils.fs.mount(
    source = "abfss://" + read_container + "@" + read_storage_account + ".dfs.core.windows.net/",
    mount_point = input_mount_point_name,
    extra_configs = configs)
  print('datalake mount done')

if str(read_directory_path) != '':
    input_file_locations = "/dbfs" + input_mount_point_name + '/' + read_directory_path
else:
    input_file_locations = "/dbfs" + input_mount_point_name
print(input_file_locations)

# COMMAND ----------

mount_point_path = input_mount_point_name + '/' + read_directory_path
print(mount_point_path)

# COMMAND ----------

'''
files_columns_dict = {}
file_abs_paths = os.path.abspath(input_file_locations)
all_files = []
for dirpath, dirnames, filenames in os.walk(file_abs_paths):
    # Walk through all files under the input_file_location
    for file in filenames:
        print(file)
        try:
            if file.endswith(('.csv', '.parquet')):
                path = os.path.join(dirpath, file)
                print(path)
                df = None
            if file.endswith('.csv'):
                df = pd.read_csv(path)

            elif file.endswith('.parquet'):
                df = pd.read_parquet(path)

            files_columns_dict[file] = df.columns
                
        except Exception as e:
            print(f"Could not process {path}: {e}")
            pass

validation_list = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in files_columns_dict.items()]))
display(validation_list)

'''

# COMMAND ----------



# This dictionary will match each file and resource set to its column names
files_columns_dict = {}

# This dictionary will determine if an asset is a file or a resource set. This distinction will be needed in the masking script. Will be converted to dataframe and added as a sheet in the configurable input file
file_directory_mapping = {'Resource_Set': [], 'File': []}

file_abs_paths = os.path.abspath(input_file_locations)

# Walk through all files and directories under the input_file_location
for dirpath, dirnames, filenames in os.walk(file_abs_paths):
    # Process each directory of files
    for dirname in dirnames:
        directory_path = os.path.join(dirpath, dirname)
        directory_relative_path = os.path.relpath(directory_path, input_file_locations)
        print(directory_relative_path)
        try:
            # Attempt to read all CSV files in the directory into a DataFrame
            csv_files = [os.path.join(directory_path, f) for f in os.listdir(directory_path) if f.endswith('.csv')]
            if csv_files:
                df = pd.concat([pd.read_csv(f) for f in csv_files], ignore_index=True)
                files_columns_dict[directory_relative_path] = df.columns.tolist()
                file_directory_mapping['Resource_Set'].append(directory_relative_path)

            # Attempt to read all Parquet files in the directory into a DataFrame
            parquet_files = [os.path.join(directory_path, f) for f in os.listdir(directory_path) if f.endswith('.parquet')]
            if parquet_files:
                df = pd.read_parquet(directory_path)  # Reading from a directory directly
                files_columns_dict[directory_relative_path] = df.columns.tolist()
                file_directory_mapping['Resource_Set'].append(directory_relative_path)

        except Exception as e:
            print(f"Could not process directory {directory_path}: {e}")

    # Process individual files within the current dirpath
    for file in filenames:
        try:
            path = os.path.join(dirpath, file)
            file_relative_path = os.path.relpath(path, input_file_locations)
            print(f' File relative path: {file_relative_path}')
            df = None
            if file.endswith('.csv'):
                df = pd.read_csv(path)
            elif file.endswith('.parquet'):
                df = pd.read_parquet(path)

            if df is not None:
                files_columns_dict[file_relative_path] = df.columns.tolist()
                file_directory_mapping['File'].append(file_relative_path)

        except Exception as e:
            print(f"Could not process file {path}: {e}")

# Convert the dictionary to a DataFrame for better visualization
validation_list = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in files_columns_dict.items()]))
display(validation_list)


# COMMAND ----------

'''
import os
import pandas as pd

# Separate dictionaries for directory and file schema mappings
resource_set_columns_dict = {}
files_columns_dict = {}
file_abs_paths = os.path.abspath(input_file_locations)

def aggregate_directory_columns(path, file_type):
    """ Helper function to read all files of a given type within a directory and return column names """
    files = [os.path.join(path, f) for f in os.listdir(path) if f.endswith(file_type)]
    if files:
        if file_type == '.csv':
            df = pd.concat([pd.read_csv(f) for f in files], ignore_index=True)
        elif file_type == '.parquet':
            df = pd.concat([pd.read_parquet(f) for f in files], ignore_index=True)
        return df.columns.tolist()
    return None

def process_directory(directory_path):
    """ Recursively process directories and files to aggregate schema information """
    local_schemas = {}
    # Process directories recursively
    for entry in os.listdir(directory_path):
        full_path = os.path.join(directory_path, entry)
        if os.path.isdir(full_path):
            # Recursive call to process subdirectories
            result = process_directory(full_path)
            if result:
                local_schemas.update(result)

    # Process files in the current directory
    for entry in os.listdir(directory_path):
        full_path = os.path.join(directory_path, entry)
        if os.path.isfile(full_path) and (entry.endswith('.csv') or entry.endswith('.parquet')):
            try:
                df = pd.read_csv(full_path) if entry.endswith('.csv') else pd.read_parquet(full_path)
                local_schemas[os.path.basename(full_path)] = df.columns.tolist()
                files_columns_dict[full_path] = df.columns.tolist()
            except Exception as e:
                print(f"Could not process file {full_path}: {e}")

    # Consolidate and check for uniform schema across this directory
    if local_schemas:
        first_schema = next(iter(local_schemas.values()))
        if all(value == first_schema for value in local_schemas.values()):
            # All schemas in this directory are the same
            return {os.path.basename(directory_path): first_schema}
        else:
            # Return individual schemas if they differ
            return {os.path.basename(full_path): schema for full_path, schema in local_schemas.items()}
    return {}

# Start processing from the root directory
consolidated_schemas = process_directory(file_abs_paths)
for k, v in consolidated_schemas.items():
    resource_set_columns_dict[k] = v

# Convert dictionaries to DataFrames
validation_list_resource_set = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in resource_set_columns_dict.items()]))
validation_list_files = pd.DataFrame(dict([(k, pd.Series(v)) for k, v in files_columns_dict.items()]))

# Display the DataFrames
print("Validation List - Resource Set:")
display(validation_list_resource_set)
print("Validation List - Files:")
display(validation_list_files)
'''

# COMMAND ----------

original_to_modified_dict = {'Original': [], 'Modified': []}
for entry in validation_list.columns:
    original_to_modified_dict['Original'].append(entry)
    original_to_modified_dict['Modified'].append(entry.replace(' ', '_').replace('-', '_'))
print(original_to_modified_dict)
original_to_modified_df = pd.DataFrame(original_to_modified_dict)
display(original_to_modified_df)
original_to_modified_df.to_csv(reference_filename_mapping, index=False)

# COMMAND ----------

validation_list.columns = validation_list.columns.str.replace(' |-', '_', regex=True)

# COMMAND ----------

print(file_directory_mapping)

# COMMAND ----------

file_directory_mapping_df = pd.DataFrame.from_dict(file_directory_mapping, orient='index').transpose()
display(file_directory_mapping_df)
#df = pd.DataFrame.from_dict(data, orient='index').transpose()

# COMMAND ----------

'''# Create a dictionary to map modified column names bacl to their original colun names - to be used to revert column names eventually
original_to_modified = {name: name.replace(' ', '_').replace('-', '_') for name in validation_list.columns}

# Remove whitespaces and dashes from the filenames so that VBA macros can still work
validation_list.columns = validation_list.columns.str.replace(' |-', '_', regex=True)
#for column in validation_list.columns:
    #print(column)
#for i in original_to_modified:
    #print(original_to_modified[i])
print(original_to_modified)
#original_to_modified_df = pd.DataFrame(original_to_modified)
#display(original_to_modified_df)
'''

# COMMAND ----------

# Create file directory_path_df, a DataFrame that references all files that can be masked, and their respective absolute paths, relative paths to the input_files location, the highest level directory they belong to after the path defined by input_files, and ultimately the file name for each.

all_paths = os.path.abspath(input_file_locations)
all_files = []
all_tables = []
all_relative_paths = []
data = []

for dirpath, dirnames, filenames in os.walk(all_paths):
    for filename in filenames:
        full_path = os.path.join(dirpath, filename)
        relative_path = os.path.relpath(dirpath, all_paths)
        highest_level_directory = relative_path.split(os.sep)[0]

        # Replace "highest level directory" with "file name" (minus extension) in instances where the table for masking is not a subdirectory that represents tables of same schema, but is instead a file directly below the directory defined as the input_path as a standalone file by itself. 
        if dirpath == all_paths:
            highest_level_directory, extension = filename.split('.')
        
        all_files.append(full_path)
        all_tables.append(highest_level_directory)
        all_relative_paths.append(relative_path)
        data.append({"full path": full_path, "relative path": relative_path, "highest level directory": highest_level_directory, "file name": filename})
    file_directory_path_df = pd.DataFrame(data)
display(file_directory_path_df)

# COMMAND ----------

# Load workbook
copyfile(masking_rules_location, local_path)
wb = load_workbook(local_path, keep_vba=True)
file_mapping_sheet = wb['file mapping']
file_directory_mapping_sheet = wb['file directory mapping']
walk_sheet = wb['walk']
sheet1 = wb['Masking Definition']

# Clear existing data in 'file mapping'
for row in file_mapping_sheet.iter_rows(min_row=1, max_row=file_mapping_sheet.max_row, min_col=1, max_col=file_mapping_sheet.max_column):
    for cell in row:
        cell.value = None

# Write validation_list dataframe to 'file mapping'
for r_idx, row in enumerate(dataframe_to_rows(validation_list, index=False, header=True), start=1):
    for c_idx, value in enumerate(row, start=1):
        file_mapping_sheet.cell(row=r_idx, column=c_idx, value=value)

# Clear existing data in 'walk'
for row in walk_sheet.iter_rows(min_row=1, max_row=walk_sheet.max_row, min_col=1, max_col=walk_sheet.max_column):
    for cell in row:
        cell.value = None

# Write validation_list dataframe to 'walk'
for r_idx, row in enumerate(dataframe_to_rows(file_directory_path_df, index=False, header=True), start=1):
    for c_idx, value in enumerate(row, start=1):
        walk_sheet.cell(row=r_idx, column=c_idx, value=value)


# Clear existing data in 'file directory mapping'
for row in file_directory_mapping_sheet.iter_rows(min_row=1, max_row=file_directory_mapping_sheet.max_row, min_col=1, max_col=file_directory_mapping_sheet.max_column):
    for cell in row:
        cell.value = None

# Write validation_list dataframe to 'file mapping'
for r_idx, row in enumerate(dataframe_to_rows(file_directory_mapping_df, index=False, header=True), start=1):
    for c_idx, value in enumerate(row, start=1):
        file_directory_mapping_sheet.cell(row=r_idx, column=c_idx, value=value)


# Create Data Validation for sheet1. Formula: Reference first row in 'file mapping' sheet
dv = DataValidation(type='list', formula1="'file mapping'!$1:$1", showDropDown=False)
sheet1.add_data_validation(dv)

# Apply data validation to column A in sheet1
for row in range(1, sheet1.max_row + 1):
    dv.add(sheet1.cell(row=row, column=1))
    #sheet1.cell(row=row, column=3).value=f'=IF(ISBLANK(A{row}), "", A{row})'

wb.save(local_path)
copyfile(local_path, configurable_masking_rulebook)

# COMMAND ----------

#Calculate and display total script runtime
end_runtime = time.time()
runtime = end_runtime - start_runtime

def seconds_to_time(seconds):
    int_seconds = int(seconds)

    # Calculate hours, minutes, and seconds
    hours = int_seconds // 3600
    int_seconds %= 3600
    minutes = int_seconds // 60
    remaining_seconds = int_seconds % 60
    fraction_seconds = seconds - int(seconds)

    formatted_time = "{:02}h:{:02}m:{:02}.{:03}s".format(hours, minutes, remaining_seconds, int(fraction_seconds*1000))

    return formatted_time
print(f'Total Runtime: {seconds_to_time(runtime)}')
